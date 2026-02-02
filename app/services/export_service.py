"""
Export service for generating Excel reports.
"""

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select, func, distinct
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.risk import Risk, NSMMapping
from app.models.action import Action
from app.models.asset import Asset
from app.models.nsm import NSMPrinciple
from app.models.ekomforskriften import EkomPrinciple, EkomMapping


class ExportService:
    """Service for exporting data to Excel."""

    # Norwegian column headers
    RISK_HEADERS = [
        "ID",
        "Tittel",
        "Beskrivelse",
        "Type",
        "Status",
        "Sannsynlighet",
        "Konsekvens",
        "Risikoscore",
        "Risikonivå",
        "Mål-sannsynlighet",
        "Mål-konsekvens",
        "Mål-risikoscore",
        "Prosjekt",
        "Eier",
        "Eksisterende kontroller",
        "Foreslåtte tiltak",
        "Neste gjennomgang",
        "Akseptert av",
        "Akseptert dato",
        "Akseptanse gyldig til",
        "Opprettet",
        "Oppdatert",
    ]

    ACTION_HEADERS = [
        "ID",
        "Tittel",
        "Beskrivelse",
        "Prioritet",
        "Status",
        "Frist",
        "Ansvarlig",
        "Avdeling",
        "Fullført dato",
        "Opprettet",
        "Oppdatert",
    ]

    ASSET_HEADERS = [
        "ID",
        "Navn",
        "Beskrivelse",
        "Type",
        "Kategori",
        "Kritikalitet",
        "Kritikalitetsnivå",
        "Lokasjon",
        "IP-adresse",
        "Serienummer",
        "Produsent",
        "Modell",
        "Netbox ID",
        "Opprettet",
        "Oppdatert",
    ]

    COVERAGE_HEADERS = [
        "Kode",
        "Kategori",
        "Tittel",
        "Versjon",
        "Antall koblede risikoer",
        "Status",
    ]

    # Colors
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    RISK_COLORS = {
        "green": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
        "yellow": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "orange": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        "red": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    }

    def __init__(self, db: AsyncSession):
        self.db = db

    def _create_workbook(self) -> Workbook:
        """Create a new workbook with basic settings."""
        wb = Workbook()
        return wb

    def _style_header_row(self, ws, headers: list[str]) -> None:
        """Apply styling to header row."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def _auto_column_width(self, ws, max_width: int = 50) -> None:
        """Auto-adjust column widths based on content."""
        for column_cells in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in column_cells
            )
            adjusted_width = min(length + 2, max_width)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

    async def export_risks(self, project_id: int | None = None) -> BytesIO:
        """
        Export risks to Excel.

        Args:
            project_id: Optional filter by project

        Returns:
            BytesIO buffer containing the Excel file
        """
        # Query risks
        query = select(Risk).options(
            selectinload(Risk.owner),
            selectinload(Risk.project),
            selectinload(Risk.accepted_by),
        )
        if project_id:
            query = query.where(Risk.project_id == project_id)
        query = query.order_by(Risk.created_at.desc())

        result = await self.db.execute(query)
        risks = result.scalars().all()

        # Create workbook
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "Risikoer"

        # Add headers
        self._style_header_row(ws, self.RISK_HEADERS)

        # Add data
        for row, risk in enumerate(risks, 2):
            ws.cell(row=row, column=1, value=risk.id)
            ws.cell(row=row, column=2, value=risk.title)
            ws.cell(row=row, column=3, value=risk.description)
            ws.cell(row=row, column=4, value=risk.risk_type.value)
            ws.cell(row=row, column=5, value=risk.status.value)
            ws.cell(row=row, column=6, value=risk.likelihood)
            ws.cell(row=row, column=7, value=risk.consequence)
            score_cell = ws.cell(row=row, column=8, value=risk.risk_score)
            ws.cell(row=row, column=9, value=risk.risk_level)
            ws.cell(row=row, column=10, value=risk.target_likelihood)
            ws.cell(row=row, column=11, value=risk.target_consequence)
            ws.cell(row=row, column=12, value=risk.target_risk_score)
            ws.cell(row=row, column=13, value=risk.project.name if risk.project else None)
            ws.cell(row=row, column=14, value=risk.owner.full_name if risk.owner else None)
            ws.cell(row=row, column=15, value=risk.existing_controls)
            ws.cell(row=row, column=16, value=risk.proposed_measures)
            ws.cell(row=row, column=17, value=risk.next_review_date.isoformat() if risk.next_review_date else None)
            ws.cell(row=row, column=18, value=risk.accepted_by.full_name if risk.accepted_by else None)
            ws.cell(row=row, column=19, value=risk.accepted_at.isoformat() if risk.accepted_at else None)
            ws.cell(row=row, column=20, value=risk.acceptance_valid_until.isoformat() if risk.acceptance_valid_until else None)
            ws.cell(row=row, column=21, value=risk.created_at.isoformat() if risk.created_at else None)
            ws.cell(row=row, column=22, value=risk.updated_at.isoformat() if risk.updated_at else None)

            # Color code risk score
            if risk.risk_color in self.RISK_COLORS:
                score_cell.fill = self.RISK_COLORS[risk.risk_color]

        self._auto_column_width(ws)

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    async def export_actions(self) -> BytesIO:
        """
        Export actions to Excel.

        Returns:
            BytesIO buffer containing the Excel file
        """
        # Query actions
        result = await self.db.execute(
            select(Action)
            .options(
                selectinload(Action.assignee),
                selectinload(Action.responsible_department),
            )
            .order_by(Action.due_date.asc().nullslast())
        )
        actions = result.scalars().all()

        # Create workbook
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "Tiltak"

        # Add headers
        self._style_header_row(ws, self.ACTION_HEADERS)

        # Add data
        for row, action in enumerate(actions, 2):
            ws.cell(row=row, column=1, value=action.id)
            ws.cell(row=row, column=2, value=action.title)
            ws.cell(row=row, column=3, value=action.description)
            ws.cell(row=row, column=4, value=action.priority.value)
            ws.cell(row=row, column=5, value=action.status.value)
            ws.cell(row=row, column=6, value=action.due_date.isoformat() if action.due_date else None)
            ws.cell(row=row, column=7, value=action.assignee.full_name if action.assignee else None)
            ws.cell(row=row, column=8, value=action.responsible_department.name if action.responsible_department else None)
            ws.cell(row=row, column=9, value=action.completed_at.isoformat() if action.completed_at else None)
            ws.cell(row=row, column=10, value=action.created_at.isoformat() if action.created_at else None)
            ws.cell(row=row, column=11, value=action.updated_at.isoformat() if action.updated_at else None)

        self._auto_column_width(ws)

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    async def export_assets(self, category: str | None = None) -> BytesIO:
        """
        Export assets to Excel.

        Args:
            category: Optional filter by category

        Returns:
            BytesIO buffer containing the Excel file
        """
        # Query assets
        query = select(Asset)
        if category:
            query = query.where(Asset.category == category)
        query = query.order_by(Asset.name)

        result = await self.db.execute(query)
        assets = result.scalars().all()

        # Create workbook
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "Assets"

        # Add headers
        self._style_header_row(ws, self.ASSET_HEADERS)

        # Add data
        for row, asset in enumerate(assets, 2):
            ws.cell(row=row, column=1, value=asset.id)
            ws.cell(row=row, column=2, value=asset.name)
            ws.cell(row=row, column=3, value=asset.description)
            ws.cell(row=row, column=4, value=asset.asset_type)
            ws.cell(row=row, column=5, value=asset.category.value)
            ws.cell(row=row, column=6, value=asset.criticality)
            ws.cell(row=row, column=7, value=asset.criticality_label)
            ws.cell(row=row, column=8, value=asset.location)
            ws.cell(row=row, column=9, value=asset.ip_address)
            ws.cell(row=row, column=10, value=asset.serial_number)
            ws.cell(row=row, column=11, value=asset.manufacturer)
            ws.cell(row=row, column=12, value=asset.model)
            ws.cell(row=row, column=13, value=asset.netbox_id)
            ws.cell(row=row, column=14, value=asset.created_at.isoformat() if asset.created_at else None)
            ws.cell(row=row, column=15, value=asset.updated_at.isoformat() if asset.updated_at else None)

        self._auto_column_width(ws)

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    async def export_coverage_report(self) -> BytesIO:
        """
        Export framework coverage report to Excel.
        Shows NSM and Ekomforskriften principles with number of linked risks.

        Returns:
            BytesIO buffer containing the Excel file
        """
        # Create workbook
        wb = self._create_workbook()

        # NSM Coverage sheet
        ws_nsm = wb.active
        ws_nsm.title = "NSM Dekning"
        self._style_header_row(ws_nsm, self.COVERAGE_HEADERS)

        # Query NSM principles with count
        nsm_result = await self.db.execute(
            select(NSMPrinciple).order_by(NSMPrinciple.sort_order)
        )
        nsm_principles = nsm_result.scalars().all()

        # Get counts per principle
        mapping_counts_result = await self.db.execute(
            select(NSMMapping.nsm_principle_id, func.count(NSMMapping.id))
            .group_by(NSMMapping.nsm_principle_id)
        )
        nsm_counts = {row[0]: row[1] for row in mapping_counts_result.all()}

        for row, principle in enumerate(nsm_principles, 2):
            count = nsm_counts.get(principle.id, 0)
            status = "Dekket" if count > 0 else "Ikke dekket"
            if principle.is_deprecated:
                status = "Utgått"

            ws_nsm.cell(row=row, column=1, value=principle.code)
            ws_nsm.cell(row=row, column=2, value=principle.category.value)
            ws_nsm.cell(row=row, column=3, value=principle.title)
            ws_nsm.cell(row=row, column=4, value=principle.version)
            ws_nsm.cell(row=row, column=5, value=count)
            ws_nsm.cell(row=row, column=6, value=status)

            # Color uncovered rows
            if count == 0 and not principle.is_deprecated:
                for col in range(1, 7):
                    ws_nsm.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                    )

        self._auto_column_width(ws_nsm)

        # Ekomforskriften Coverage sheet
        ws_ekom = wb.create_sheet(title="Ekomforskriften Dekning")
        self._style_header_row(ws_ekom, self.COVERAGE_HEADERS)

        # Query Ekomforskriften principles with count
        ekom_result = await self.db.execute(
            select(EkomPrinciple).order_by(EkomPrinciple.sort_order)
        )
        ekom_principles = ekom_result.scalars().all()

        # Get counts per principle
        ekom_mapping_counts_result = await self.db.execute(
            select(EkomMapping.ekom_principle_id, func.count(EkomMapping.id))
            .group_by(EkomMapping.ekom_principle_id)
        )
        ekom_counts = {row[0]: row[1] for row in ekom_mapping_counts_result.all()}

        for row, principle in enumerate(ekom_principles, 2):
            count = ekom_counts.get(principle.id, 0)
            status = "Dekket" if count > 0 else "Ikke dekket"
            if principle.is_deprecated:
                status = "Utgått"

            ws_ekom.cell(row=row, column=1, value=f"§ {principle.code}")
            ws_ekom.cell(row=row, column=2, value=principle.category.value)
            ws_ekom.cell(row=row, column=3, value=principle.title)
            ws_ekom.cell(row=row, column=4, value=principle.version)
            ws_ekom.cell(row=row, column=5, value=count)
            ws_ekom.cell(row=row, column=6, value=status)

            # Color uncovered rows
            if count == 0 and not principle.is_deprecated:
                for col in range(1, 7):
                    ws_ekom.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                    )

        self._auto_column_width(ws_ekom)

        # Summary sheet
        ws_summary = wb.create_sheet(title="Sammendrag")

        # NSM summary
        nsm_total = len(nsm_principles)
        nsm_covered = len([p for p in nsm_principles if nsm_counts.get(p.id, 0) > 0])
        nsm_deprecated = len([p for p in nsm_principles if p.is_deprecated])
        nsm_active = nsm_total - nsm_deprecated
        nsm_coverage_pct = round(nsm_covered / nsm_active * 100, 1) if nsm_active > 0 else 0

        ws_summary.cell(row=1, column=1, value="NSM Grunnprinsipper")
        ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws_summary.cell(row=2, column=1, value="Totalt antall prinsipper:")
        ws_summary.cell(row=2, column=2, value=nsm_total)
        ws_summary.cell(row=3, column=1, value="Aktive prinsipper:")
        ws_summary.cell(row=3, column=2, value=nsm_active)
        ws_summary.cell(row=4, column=1, value="Utgåtte prinsipper:")
        ws_summary.cell(row=4, column=2, value=nsm_deprecated)
        ws_summary.cell(row=5, column=1, value="Prinsipper med dekning:")
        ws_summary.cell(row=5, column=2, value=nsm_covered)
        ws_summary.cell(row=6, column=1, value="Dekningsgrad:")
        ws_summary.cell(row=6, column=2, value=f"{nsm_coverage_pct}%")

        # Ekomforskriften summary
        ekom_total = len(ekom_principles)
        ekom_covered = len([p for p in ekom_principles if ekom_counts.get(p.id, 0) > 0])
        ekom_deprecated = len([p for p in ekom_principles if p.is_deprecated])
        ekom_active = ekom_total - ekom_deprecated
        ekom_coverage_pct = round(ekom_covered / ekom_active * 100, 1) if ekom_active > 0 else 0

        ws_summary.cell(row=8, column=1, value="Ekomforskriften")
        ws_summary.cell(row=8, column=1).font = Font(bold=True, size=14)
        ws_summary.cell(row=9, column=1, value="Totalt antall paragrafer:")
        ws_summary.cell(row=9, column=2, value=ekom_total)
        ws_summary.cell(row=10, column=1, value="Aktive paragrafer:")
        ws_summary.cell(row=10, column=2, value=ekom_active)
        ws_summary.cell(row=11, column=1, value="Utgåtte paragrafer:")
        ws_summary.cell(row=11, column=2, value=ekom_deprecated)
        ws_summary.cell(row=12, column=1, value="Paragrafer med dekning:")
        ws_summary.cell(row=12, column=2, value=ekom_covered)
        ws_summary.cell(row=13, column=1, value="Dekningsgrad:")
        ws_summary.cell(row=13, column=2, value=f"{ekom_coverage_pct}%")

        # Report date
        ws_summary.cell(row=15, column=1, value="Rapport generert:")
        ws_summary.cell(row=15, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

        self._auto_column_width(ws_summary)

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
